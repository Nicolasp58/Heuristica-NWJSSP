"""
main2.py
========
Orquestador principal – Trabajo 2, Heurística EAFIT.
Búsqueda local PURA (sin Multi-Start).

Distribución del tiempo
-----------------------
El tiempo total (1 hora) se distribuye proporcionalmente al peso n*m de
cada instancia. Ejecución secuencial con deadlines absolutos.

Formato Excel (Anexo 3):
  A1 = Z,  B1 = tiempo_ms
  Fila 2 = s_0, s_1, ..., s_{n-1}

Solución inicial: Noise(r=0.5, nsol=1).

Orden de instancias: debe coincidir exactamente con el orden de lb.txt.
"""

import os
import time
from openpyxl import Workbook

from Read        import read_nwjssp
from Noise       import Noise
from LocalSearch import run_local_search, NEIGHBORHOODS, CRITERIA

# ── Parámetros ────────────────────────────────────────────────────────────────
NOISE_R    = 0.5
NOISE_NSOL = 1

INSTANCES_FOLDER = "NWJSSP Instances"
LB_FILE          = "lb.txt"
TOTAL_TIME_S     = 3_600
STUDENT_NAME     = "NicolasPena"

# Orden exacto — debe coincidir línea a línea con lb.txt
INSTANCE_ORDER = [
    "ft06.txt",
    "ft06r.txt",
    "ft10.txt",
    "ft10r.txt",
    "ft20.txt",
    "ft20r.txt",
    "tai_j10_m10_1.txt",
    "tai_j10_m10_1r.txt",
    "tai_j100_m10_1.txt",
    "tai_j100_m10_1r.txt",
    "tai_j100_m100_1.txt",
    "tai_j100_m100_1r.txt",
    "tai_j1000_m10_1.txt",
    "tai_j1000_m10_1r.txt",
    "tai_j1000_m100_1.txt",
    "tai_j1000_m100_1r.txt",
    "tai_j1000_m1000_1.txt",
    "tai_j1000_m1000_1r.txt",
]


def load_lower_bounds(lb_path):
    with open(lb_path) as f:
        return [int(line.strip()) for line in f if line.strip()]


def _instance_weight(n, m):
    return n * m


def _write_sheet_annex3(ws, Z, t_ms, start_times, n):
    ws.cell(row=1, column=1, value=Z)
    ws.cell(row=1, column=2, value=round(t_ms))
    for j in range(n):
        ws.cell(row=2, column=j + 1, value=start_times[j])


def _build_excel(neighborhood, criterion, results_dict, instance_names):
    fname = f"NWJSSP_{STUDENT_NAME}_LS_{neighborhood}_{criterion}.xlsx"
    wb    = Workbook()
    wb.remove(wb.active)
    for iname in instance_names:
        if iname not in results_dict:
            continue
        rec = results_dict[iname]
        ws  = wb.create_sheet(title=iname[:31])
        _write_sheet_annex3(ws, rec['Z'], rec['time_ms'], rec['start_times'], rec['n'])
    wb.save(fname)
    print(f"  -> {fname}")


def main():
    print("=" * 60)
    print("  NWJSSP — Búsqueda Local Pura (Trabajo 2)")
    print(f"  {STUDENT_NAME} — Universidad EAFIT")
    print("=" * 60)

    if not os.path.isdir(INSTANCES_FOLDER):
        print(f"ERROR: No se encontro '{INSTANCES_FOLDER}/'."); return
    if not os.path.isfile(LB_FILE):
        print(f"ERROR: No se encontro '{LB_FILE}'."); return

    lb_list = load_lower_bounds(LB_FILE)

    # Verificar qué instancias están disponibles
    available = set(os.listdir(INSTANCES_FOLDER))
    instance_files = []
    for fname in INSTANCE_ORDER:
        if fname in available:
            instance_files.append(fname)
        else:
            print(f"  ADVERTENCIA: '{fname}' no encontrado en '{INSTANCES_FOLDER}/' — se omite.")

    n_inst = len(instance_files)
    if n_inst == 0:
        print("ERROR: No hay instancias disponibles."); return

    # Los LB corresponden al índice en INSTANCE_ORDER (no en instance_files filtrado)
    # Construir mapa nombre -> LB
    lb_map = {}
    for i, fname in enumerate(INSTANCE_ORDER):
        iname = fname.replace(".txt", "")
        if i < len(lb_list):
            lb_map[iname] = lb_list[i]

    # Leer dimensiones
    print("\n  Distribución de tiempo (proporcional a n*m):")
    print(f"  {'Instancia':<28} {'n':>6} {'m':>6} {'peso':>12} {'tiempo(s)':>10} {'LB':>14}")
    print(f"  {'─'*28} {'─'*6} {'─'*6} {'─'*12} {'─'*10} {'─'*14}")

    instance_dims = []
    for fname in instance_files:
        fp = os.path.join(INSTANCES_FOLDER, fname)
        n, m, _, _ = read_nwjssp(fp)
        instance_dims.append((n, m))

    weights      = [_instance_weight(n, m) for n, m in instance_dims]
    total_weight = sum(weights)
    n_combos     = len(NEIGHBORHOODS) * len(CRITERIA)   # 6
    time_per_inst = [TOTAL_TIME_S * w / total_weight for w in weights]

    for fname, (n, m), t in zip(instance_files, instance_dims, time_per_inst):
        iname = fname.replace(".txt", "")
        lb    = lb_map.get(iname, 0)
        print(f"  {iname:<28} {n:>6} {m:>6} {n*m:>12,} {t:>10.1f}s {lb:>14,}")

    print(f"\n  Total: {sum(time_per_inst):.1f}s  |  {n_combos} combos/instancia")
    print(f"  Instancias disponibles: {n_inst} / {len(INSTANCE_ORDER)}\n")

    all_results    = {f"{nbh}_{crit}": {} for nbh in NEIGHBORHOODS for crit in CRITERIA}
    instance_names = []
    t_start        = time.time()

    for idx, filename in enumerate(instance_files):
        iname    = filename.replace(".txt", "")
        filepath = os.path.join(INSTANCES_FOLDER, filename)
        lb       = lb_map.get(iname, 0)
        n, m     = instance_dims[idx]

        elapsed_before = sum(time_per_inst[:idx])
        secs_per_combo = time_per_inst[idx] / n_combos

        print(f"\n{'─'*60}")
        print(f"  [{idx+1}/{n_inst}] {iname}  (LB={lb:,})")
        print(f"  n={n}, m={m}  |  {time_per_inst[idx]:.1f}s  ({secs_per_combo:.1f}s/combo)")
        print(f"{'─'*60}")

        n2, m2, ops, rd = read_nwjssp(filepath)
        Z_init, seq_init, _, _, t_c = Noise(n2, m2, ops, rd, NOISE_R, NOISE_NSOL)
        gap_c = round((Z_init - lb) / lb * 100, 2) if lb > 0 else 0
        print(f"  Noise(r={NOISE_R}, nsol={NOISE_NSOL}): Z={Z_init:,}  gap={gap_c}%  t={round(t_c)}ms")

        instance_names.append(iname)
        combo_idx = 0

        for nbh in NEIGHBORHOODS:
            for crit in CRITERIA:
                combo_start_offset = elapsed_before + combo_idx * secs_per_combo
                deadline = t_start + combo_start_offset + secs_per_combo
                combo_idx += 1

                now = time.time()
                if deadline <= now:
                    deadline = now + 2.0

                res = run_local_search(
                    n2, m2, ops, rd,
                    seq_init, Z_init,
                    nbh, crit,
                    deadline=deadline
                )
                gap = round((res['Z'] - lb) / lb * 100, 2) if lb > 0 else 0
                print(f"  {nbh:6s}-{crit}: Z={res['Z']:,}  gap={gap}%  "
                      f"iters={res['iterations']}  evals={res['evaluations']}  "
                      f"t={round(res['time_ms'])}ms")

                all_results[f"{nbh}_{crit}"][iname] = {**res, 'n': n2}

    print(f"\n{'='*60}")
    print("  Generando archivos Excel (formato Anexo 3)...")
    for nbh in NEIGHBORHOODS:
        for crit in CRITERIA:
            _build_excel(nbh, crit, all_results[f"{nbh}_{crit}"], instance_names)

    elapsed_min = (time.time() - t_start) / 60
    print(f"\n  Tiempo total: {elapsed_min:.2f} minutos")
    print("  Listo!")


if __name__ == "__main__":
    main()