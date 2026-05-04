import os
import time
from openpyxl import Workbook

from Read  import read_nwjssp
from Noise import Noise
from VND   import run_vnd
from ILS   import run_ils

# ══════════════════════════════════════════════════════════════════════════════
INSTANCES_FOLDER = "NWJSSP Instances"
LB_FILE          = "lb.txt"
STUDENT_NAME     = "NicolasPena"

TOTAL_TIME_VND_S = 3_600
TOTAL_TIME_ILS_S = 3_600
MIN_TIME_S       = 30

NOISE_R    = 0.5
NOISE_NSOL = 1
PERTURB_K  = 4

INSTANCE_ORDER = [
    "ft06.txt",        "ft06r.txt",
    "ft10.txt",        "ft10r.txt",
    "ft20.txt",        "ft20r.txt",
    "tai_j10_m10_1.txt",   "tai_j10_m10_1r.txt",
    "tai_j100_m10_1.txt",  "tai_j100_m10_1r.txt",
    "tai_j100_m100_1.txt", "tai_j100_m100_1r.txt",
    "tai_j1000_m10_1.txt", "tai_j1000_m10_1r.txt",
    "tai_j1000_m100_1.txt","tai_j1000_m100_1r.txt",
    "tai_j1000_m1000_1.txt","tai_j1000_m1000_1r.txt",
]
# ══════════════════════════════════════════════════════════════════════════════


def load_lb_map(lb_path):
    with open(lb_path) as f:
        vals = [int(l.strip()) for l in f if l.strip()]
    return {INSTANCE_ORDER[i].replace('.txt',''): vals[i]
            for i in range(min(len(INSTANCE_ORDER), len(vals)))}


def next_instance_time(pool_s, remaining_dims, min_s=MIN_TIME_S):
    """
    Calcula el tiempo para la próxima instancia (remaining_dims[0])
    usando el pool de tiempo disponible restante.
    """
    n_left = len(remaining_dims)
    if n_left == 0:
        return 0.0
    min_future  = min_s * (n_left - 1)
    available   = max(min_s, pool_s - min_future)
    weights     = [ni * mi for ni, mi in remaining_dims]
    total_w     = sum(weights) or 1
    excess      = max(0.0, pool_s - min_s * n_left)
    my_time     = min_s + excess * weights[0] / total_w
    return min(my_time, available)


def _write_sheet(ws, Z, t_ms, start_times, n):
    """Formato Anexo 3: fila1=[Z, t_ms]  fila2=[s0..s_{n-1}]"""
    ws.cell(row=1, column=1, value=Z)
    ws.cell(row=1, column=2, value=round(t_ms))
    for j in range(n):
        ws.cell(row=2, column=j + 1, value=start_times[j])


def _build_excel(method_name, results_dict, instance_names):
    fname = f"NWJSSP_{STUDENT_NAME}_{method_name}.xlsx"
    wb    = Workbook()
    wb.remove(wb.active)
    for iname in instance_names:
        if iname not in results_dict:
            continue
        rec = results_dict[iname]
        ws  = wb.create_sheet(title=iname[:31])
        _write_sheet(ws, rec['Z'], rec['time_ms'], rec['start_times'], rec['n'])
    wb.save(fname)
    print(f"  -> {fname}")


# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 62)
    print("  NWJSSP — VND + ILS (Trabajo 3)")
    print(f"  {STUDENT_NAME} — Universidad EAFIT")
    print("=" * 62)
    print(f"  VND : FI, swap→insert(k=3)→2-opt(len=3)")
    print(f"  ILS : inicio=VND, FI interno, SA(T0=0.02), k={PERTURB_K}")
    print(f"  Tiempo: {TOTAL_TIME_VND_S}s VND + {TOTAL_TIME_ILS_S}s ILS")
    print(f"  Pool dinámico | mín {MIN_TIME_S}s/instancia | sobrante redistribuido")

    if not os.path.isdir(INSTANCES_FOLDER):
        print(f"ERROR: No se encontró '{INSTANCES_FOLDER}/'"); return
    if not os.path.isfile(LB_FILE):
        print(f"ERROR: No se encontró '{LB_FILE}'"); return

    lb_map = load_lb_map(LB_FILE)
    available = set(os.listdir(INSTANCES_FOLDER))
    instance_files = [f for f in INSTANCE_ORDER if f in available]

    if not instance_files:
        print("ERROR: Sin instancias."); return

    dims = []
    for fname in instance_files:
        nv, mv, _, _ = read_nwjssp(os.path.join(INSTANCES_FOLDER, fname))
        dims.append((nv, mv))

    # tabla de tiempos estimados (referencia)
    def _estimate_times(total_s, dims_list, min_s=MIN_TIME_S):
        n         = len(dims_list)
        floor     = min_s * n
        remaining = max(0.0, total_s - floor)
        weights   = [ni * mi for ni, mi in dims_list]
        total_w   = sum(weights) or 1
        return [min_s + remaining * w / total_w for w in weights]


    tvnd_est = _estimate_times(TOTAL_TIME_VND_S, dims)
    tils_est = _estimate_times(TOTAL_TIME_ILS_S, dims)
    print(f"\n  {'Instancia':<30} {'n':>6} {'m':>6} {'VND_est':>10} {'ILS_est':>10}")
    print(f"  {'─'*30} {'─'*6} {'─'*6} {'─'*10} {'─'*10}")
    for i, fname in enumerate(instance_files):
        iname = fname.replace('.txt','')
        n, m  = dims[i]
        print(f"  {iname:<30} {n:>6} {m:>6} {tvnd_est[i]:>10.1f} {tils_est[i]:>10.1f}")
    print()

    # leer instancias + Noise
    all_data  = {}
    all_noise = {}
    instance_names = []
    for fname in instance_files:
        iname = fname.replace('.txt','')
        n, m, ops, rd = read_nwjssp(os.path.join(INSTANCES_FOLDER, fname))
        Z0, seq0, _, _, _ = Noise(n, m, ops, rd, NOISE_R, NOISE_NSOL)
        all_data[iname]  = (n, m, ops, rd)
        all_noise[iname] = (Z0, seq0)
        instance_names.append(iname)

    results_vnd = {}
    results_ils = {}

    # ══════════════════════════════════════════════
    #  BLOQUE 1: VND con pool dinámico
    # ══════════════════════════════════════════════
    print("=" * 62)
    print("  BLOQUE 1/2 — VND")
    print("=" * 62)

    pool_vnd = float(TOTAL_TIME_VND_S)

    for idx, fname in enumerate(instance_files):
        iname         = fname.replace('.txt','')
        n, m, ops, rd = all_data[iname]
        lb            = lb_map.get(iname, 0)
        Z0, seq0      = all_noise[iname]

        remaining_dims = dims[idx:]
        t_assigned = max(next_instance_time(pool_vnd, remaining_dims), 1.0)
        deadline   = time.time() + t_assigned

        print(f"\n  [{idx+1}/{len(instance_files)}] {iname}"
              f"  (LB={lb:,}  pool={pool_vnd:.1f}s  t_asignado={t_assigned:.1f}s)")

        t_start = time.time()
        res = run_vnd(n, m, ops, rd, seq0, Z0, deadline=deadline)
        t_used = time.time() - t_start

        pool_vnd -= t_used
        pool_vnd  = max(0.0, pool_vnd)

        g   = round((res['Z']-lb)/lb*100, 2) if lb else 0
        stp = 'tiempo' if res['time_out'] else 'optimo_local'
        print(f"    Z={res['Z']:,}  gap={g}%  iters={res['iterations']}"
              f"  t_usado={round(t_used,2)}s  sobrante={round(t_assigned-t_used,2)}s"
              f"  [{stp}]  pool_restante={pool_vnd:.1f}s")

        results_vnd[iname] = {**res, 'n': n}

    # ══════════════════════════════════════════════
    #  BLOQUE 2: ILS con pool dinámico
    # ══════════════════════════════════════════════
    print()
    print("=" * 62)
    print("  BLOQUE 2/2 — ILS")
    print("=" * 62)

    pool_ils = float(TOTAL_TIME_ILS_S)

    for idx, fname in enumerate(instance_files):
        iname         = fname.replace('.txt','')
        n, m, ops, rd = all_data[iname]
        lb            = lb_map.get(iname, 0)
        res_vnd       = results_vnd[iname]

        remaining_dims = dims[idx:]
        t_assigned = max(next_instance_time(pool_ils, remaining_dims), 1.0)
        deadline   = time.time() + t_assigned

        print(f"\n  [{idx+1}/{len(instance_files)}] {iname}"
              f"  (LB={lb:,}  pool={pool_ils:.1f}s  t_asignado={t_assigned:.1f}s)")
        print(f"    Inicio VND: Z={res_vnd['Z']:,}")

        t_start = time.time()
        t_conv_s = results_vnd[iname]['time_ms'] / 1000.0
        res = run_ils(n, m, ops, rd,
                      res_vnd['sequence'], res_vnd['Z'],
                      perturb_k=PERTURB_K, t_conv_s=t_conv_s, deadline=deadline)
        t_used = time.time() - t_start

        pool_ils -= t_used
        pool_ils  = max(0.0, pool_ils)

        g   = round((res['Z']-lb)/lb*100, 2) if lb else 0
        stp = 'tiempo' if res['time_out'] else 'convergencia'
        print(f"    Z={res['Z']:,}  gap={g}%  ILS-iters={res['ils_iters']}"
              f"  acep_peor={res['accepted_worse']}"
              f"  vnd_init={res['vnd_init_t']}s  vnd_int={res['vnd_int_t']}s"
              f"  t_usado={round(t_used,2)}s  sobrante={round(t_assigned-t_used,2)}s"
              f"  [{stp}]  pool_restante={pool_ils:.1f}s")

        results_ils[iname] = {**res, 'n': n}

    # ── Excel ──────────────────────────────────────────────────────────────────
    print()
    print("=" * 62)
    print("  Generando Excel (Anexo 3)...")
    _build_excel("VND", results_vnd, instance_names)
    _build_excel("ILS", results_ils, instance_names)
    print("  Listo.")


if __name__ == "__main__":
    main()