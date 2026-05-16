import os
import time
from openpyxl import Workbook

from Read      import read_nwjssp
from Noise     import Noise
from GAVND     import run_gavnd
from BRKGA_ILS import run_brkga_ils

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN GENERAL
# ══════════════════════════════════════════════════════════════════════════════

INSTANCES_FOLDER = "NWJSSP Instances"
LB_FILE          = "lb.txt"
STUDENT_NAME     = "NicolasPena"

TOTAL_TIME_GAVND_S    = 3_600   # 1 hora para GAVND
TOTAL_TIME_BRKGA_ILS_S = 3_600  # 1 hora para BRKGA+ILS

INSTANCE_ORDER = [
    "ft06.txt",              "ft06r.txt",
    "ft10.txt",              "ft10r.txt",
    "ft20.txt",              "ft20r.txt",
    "tai_j10_m10_1.txt",     "tai_j10_m10_1r.txt",
    "tai_j100_m10_1.txt",    "tai_j100_m10_1r.txt",
    "tai_j100_m100_1.txt",   "tai_j100_m100_1r.txt",
    "tai_j1000_m10_1.txt",   "tai_j1000_m10_1r.txt",
    "tai_j1000_m100_1.txt",  "tai_j1000_m100_1r.txt",
    "tai_j1000_m1000_1.txt", "tai_j1000_m1000_1r.txt",
]

# ══════════════════════════════════════════════════════════════════════════════
#  PARÁMETROS
# ══════════════════════════════════════════════════════════════════════════════

# ── GAVND ─────────────────────────────────────────────────────────────────────
GAVND_POP_SIZE     = 70
GAVND_P_MUT        = 0.20
GAVND_TOURNAMENT_K = 4
GAVND_T_VND_MAX    = 1.0
GAVND_VND_EVERY_N  = 1

# ── BRKGA+ILS ─────────────────────────────────────────────────────────────────
BRKGA_POP_SIZE      = 100
BRKGA_P_ELITE       = 0.10
BRKGA_P_MUTANT      = 0.10
BRKGA_P_BIAS        = 0.70
BRKGA_ILS_PERTURB_K = 4
BRKGA_T_ILS_MAX     = 1.5
BRKGA_ILS_EVERY_N   = 1

# ── Tiempos mínimos por tamaño de instancia ────────────────────────────────────
MIN_TIME_BY_SIZE = [
    (100,          15),    
    (1_000,       150),    
    (10_000,      180),    
    (float('inf'), 180),   
]

NOISE_R    = 0.5
NOISE_NSOL = 1


# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def load_lb_map(lb_path):
    with open(lb_path) as f:
        vals = [int(l.strip()) for l in f if l.strip()]
    return {INSTANCE_ORDER[i].replace('.txt', ''): vals[i]
            for i in range(min(len(INSTANCE_ORDER), len(vals)))}


def min_time_for(n, m):
    nm = n * m
    for threshold, t in MIN_TIME_BY_SIZE:
        if nm <= threshold:
            return t
    return MIN_TIME_BY_SIZE[-1][1]


def next_instance_time(pool_s, remaining_dims):
    """
    Calcula el tiempo asignado a la próxima instancia (remaining_dims[0])
    dado el pool disponible. Garantiza el mínimo de cada instancia futura
    y distribuye el excedente proporcionalmente a n×m.
    """
    n_left = len(remaining_dims)
    if n_left == 0:
        return 0.0
    mins_future = sum(min_time_for(n, m) for n, m in remaining_dims[1:])
    available   = max(min_time_for(*remaining_dims[0]), pool_s - mins_future)
    weights     = [n * m for n, m in remaining_dims]
    total_w     = sum(weights) or 1
    excess      = max(0.0, pool_s - sum(min_time_for(n, m) for n, m in remaining_dims))
    my_time     = min_time_for(*remaining_dims[0]) + excess * weights[0] / total_w
    return min(my_time, available)


def _write_sheet_annex3(ws, Z, t_ms, start_times, n):
    """Formato Anexo 3 exacto: fila1=[Z, t_ms], fila2=[s_0..s_{n-1}]"""
    ws.cell(row=1, column=1, value=Z)
    ws.cell(row=1, column=2, value=round(t_ms))
    for j in range(n):
        ws.cell(row=2, column=j + 1, value=start_times[j])


def _build_excel(method_name, results_dict, instance_names):
    fname = f"NWJSSP_{STUDENT_NAME}_{method_name}.xlsx"
    wb    = Workbook()
    wb.remove(wb.active)
    for iname in instance_names:
        if iname not in results_dict:
            continue
        rec = results_dict[iname]
        ws  = wb.create_sheet(title=iname[:31])
        _write_sheet_annex3(ws, rec['Z'], rec['time_ms'], rec['start_times'], rec['n'])
    wb.save(fname)
    print(f"  -> {fname}")


def _print_time_table(total_s, label, inst_files, dims, lb_map):
    """Imprime tabla estimada de distribución de tiempo."""
    t_est = []
    pool  = float(total_s)
    for idx in range(len(dims)):
        t = next_instance_time(pool, dims[idx:])
        t_est.append(t)
        pool = max(0.0, pool - t)

    print(f"\n  Distribución de tiempo estimada — {label}")
    print(f"  {'Instancia':<30} {'n':>6} {'m':>6} {'t_min':>7} {'t_est':>8} {'LB':>18}")
    print(f"  {'─'*30} {'─'*6} {'─'*6} {'─'*7} {'─'*8} {'─'*18}")
    for i, fname in enumerate(inst_files):
        iname  = fname.replace('.txt', '')
        ni, mi = dims[i]
        lb     = lb_map.get(iname, 0)
        print(f"  {iname:<30} {ni:>6} {mi:>6} "
              f"{min_time_for(ni,mi):>7.0f} {t_est[i]:>8.1f} {lb:>18,}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 68)
    print("  NWJSSP — Algoritmos Evolutivos (Trabajo 4)")
    print(f"  {STUDENT_NAME} — Universidad EAFIT")
    print("=" * 68)
    print(f"\n  GAVND    : pop={GAVND_POP_SIZE}, p_mut={GAVND_P_MUT}, "
          f"k={GAVND_TOURNAMENT_K}, t_vnd={GAVND_T_VND_MAX}s, "
          f"vnd_every={GAVND_VND_EVERY_N}")
    print(f"  BRKGA+ILS: pop={BRKGA_POP_SIZE}, elite={BRKGA_P_ELITE}, "
          f"mut={BRKGA_P_MUTANT}, bias={BRKGA_P_BIAS}, "
          f"k_ils={BRKGA_ILS_PERTURB_K}, t_ils={BRKGA_T_ILS_MAX}s, "
          f"ils_every={BRKGA_ILS_EVERY_N}")

    if not os.path.isdir(INSTANCES_FOLDER):
        print(f"\nERROR: No se encontró '{INSTANCES_FOLDER}/'"); return
    if not os.path.isfile(LB_FILE):
        print(f"\nERROR: No se encontró '{LB_FILE}'"); return

    lb_map     = load_lb_map(LB_FILE)
    available  = set(os.listdir(INSTANCES_FOLDER))
    inst_files = [f for f in INSTANCE_ORDER if f in available]

    if not inst_files:
        print("\nERROR: Sin instancias."); return

    # Leer dimensiones
    dims = []
    for fname in inst_files:
        nv, mv, _, _ = read_nwjssp(os.path.join(INSTANCES_FOLDER, fname))
        dims.append((nv, mv))

    _print_time_table(TOTAL_TIME_GAVND_S,     "GAVND",     inst_files, dims, lb_map)
    _print_time_table(TOTAL_TIME_BRKGA_ILS_S, "BRKGA+ILS", inst_files, dims, lb_map)

    # ── Precargar instancias + solución Noise inicial ──────────────────────────
    print("\n" + "=" * 68)
    print("  Precargando instancias...")
    print("=" * 68)

    all_data  = {}
    all_noise = {}
    inames    = []
    for fname in inst_files:
        iname = fname.replace('.txt', '')
        n, m, ops, rd = read_nwjssp(os.path.join(INSTANCES_FOLDER, fname))
        Z0, seq0, _, _, _ = Noise(n, m, ops, rd, NOISE_R, NOISE_NSOL)
        lb = lb_map.get(iname, 0)
        all_data[iname]  = (n, m, ops, rd)
        all_noise[iname] = (Z0, seq0)
        inames.append(iname)
        gap0 = round((Z0 - lb) / lb * 100, 2) if lb else 0
        print(f"  {iname:<30}  Z_noise={Z0:>16,}  gap={gap0:.2f}%")

    results_gavnd    = {}
    results_brkga    = {}

    # ══════════════════════════════════════════════════════════════════════
    #  BLOQUE 1: GAVND
    # ══════════════════════════════════════════════════════════════════════
    print("\n" + "=" * 68)
    print("  BLOQUE 1/2 — GAVND (GA + VND)")
    print("=" * 68)
    pool_gavnd = float(TOTAL_TIME_GAVND_S)

    for idx, fname in enumerate(inst_files):
        iname         = fname.replace('.txt', '')
        n, m, ops, rd = all_data[iname]
        lb            = lb_map.get(iname, 0)
        Z0, _         = all_noise[iname]

        t_assigned = max(next_instance_time(pool_gavnd, dims[idx:]), 1.0)
        deadline   = time.time() + t_assigned

        gap0 = round((Z0 - lb) / lb * 100, 2) if lb else 0
        print(f"\n  [{idx+1}/{len(inst_files)}] {iname}  "
              f"(LB={lb:,}  pool={pool_gavnd:.0f}s  t={t_assigned:.1f}s)")
        print(f"    Noise inicial: Z={Z0:,}  gap={gap0:.2f}%")

        t_inst = time.time()
        res = run_gavnd(
            n, m, ops, rd,
            pop_size    =GAVND_POP_SIZE,
            p_mut       =GAVND_P_MUT,
            tournament_k=GAVND_TOURNAMENT_K,
            t_vnd_max   =GAVND_T_VND_MAX,
            vnd_every_n =GAVND_VND_EVERY_N,
            deadline    =deadline,
        )
        t_used     = time.time() - t_inst
        pool_gavnd = max(0.0, pool_gavnd - t_used)

        gap_f  = round((res['Z'] - lb) / lb * 100, 2) if lb else 0
        mejora = round((Z0 - res['Z']) / Z0 * 100, 2) if Z0 else 0
        print(f"    Z={res['Z']:,}  gap={gap_f:.2f}%  mejora={mejora:.2f}%  "
              f"gen={res['generations']}  vnd={res['vnd_calls']}  "
              f"t={t_used:.1f}s  pool_rest={pool_gavnd:.0f}s")

        results_gavnd[iname] = {**res, 'n': n}

    # ══════════════════════════════════════════════════════════════════════
    #  BLOQUE 2: BRKGA+ILS
    # ══════════════════════════════════════════════════════════════════════
    print("\n" + "=" * 68)
    print("  BLOQUE 2/2 — BRKGA+ILS")
    print("=" * 68)
    pool_brkga = float(TOTAL_TIME_BRKGA_ILS_S)

    for idx, fname in enumerate(inst_files):
        iname         = fname.replace('.txt', '')
        n, m, ops, rd = all_data[iname]
        lb            = lb_map.get(iname, 0)
        Z0, _         = all_noise[iname]

        t_assigned = max(next_instance_time(pool_brkga, dims[idx:]), 1.0)
        deadline   = time.time() + t_assigned

        gap0 = round((Z0 - lb) / lb * 100, 2) if lb else 0
        print(f"\n  [{idx+1}/{len(inst_files)}] {iname}  "
              f"(LB={lb:,}  pool={pool_brkga:.0f}s  t={t_assigned:.1f}s)")
        print(f"    Noise inicial: Z={Z0:,}  gap={gap0:.2f}%")

        t_inst = time.time()
        res = run_brkga_ils(
            n, m, ops, rd,
            pop_size     =BRKGA_POP_SIZE,
            p_elite      =BRKGA_P_ELITE,
            p_mutant     =BRKGA_P_MUTANT,
            p_bias       =BRKGA_P_BIAS,
            ils_perturb_k=BRKGA_ILS_PERTURB_K,
            t_ils_max    =BRKGA_T_ILS_MAX,
            ils_every_n  =BRKGA_ILS_EVERY_N,
            deadline     =deadline,
        )
        t_used     = time.time() - t_inst
        pool_brkga = max(0.0, pool_brkga - t_used)

        gap_f  = round((res['Z'] - lb) / lb * 100, 2) if lb else 0
        mejora = round((Z0 - res['Z']) / Z0 * 100, 2) if Z0 else 0
        ils_eff = (f"{res['ils_improved']}/{res['ils_calls']}"
                   if res['ils_calls'] > 0 else "0/0")
        print(f"    Z={res['Z']:,}  gap={gap_f:.2f}%  mejora={mejora:.2f}%  "
              f"gen={res['generations']}  ils={ils_eff}  "
              f"t={t_used:.1f}s  pool_rest={pool_brkga:.0f}s")

        results_brkga[iname] = {**res, 'n': n}

    # ── Excel (Anexo 3) ────────────────────────────────────────────────────────
    print("\n" + "=" * 68)
    print("  Generando Excel (Anexo 3)...")
    _build_excel("GAVND",    results_gavnd, inames)
    _build_excel("BRKGA_ILS", results_brkga, inames)
    print("  Listo.")


if __name__ == "__main__":
    main()