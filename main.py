"""
main.py
-------
Orquestador principal del proyecto NWJSSP.
Universidad EAFIT — Curso: Heuristica
Estudiante: Nicolas Pena

Ejecuta los tres algoritmos (Constructivo, GRASP, Noise) sobre todas las
instancias .txt disponibles en la carpeta de instancias y genera archivos
Excel con los resultados en el formato requerido (Anexo 3).

============================================================
PARAMETROS DEL ALGORITMO — modificar solo aqui
============================================================
"""

import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from Read import read_nwjssp
from Evaluator import precompute, lower_bound
from Constructive import Constructive
from GRASP1 import GRASP1
from Noise import Noise

# ============================================================
# PARAMETROS — editar estos valores para cambiar el comportamiento
# ============================================================
INSTANCES_FOLDER = "NWJSSP Instances"   # Carpeta con los archivos .txt de instancias

nsol  = 100    # Numero de soluciones (iteraciones de GRASP y Noise)
alpha = 0.05   # Parametro GRASP: controla tamano de la RCL (0=greedy, 1=aleatorio)
K     = 5      # Tamano maximo de la RCL (referencia conceptual; se usa alpha como criterio)
r     = 5      # Amplitud del ruido para el Noising Method
# ============================================================


def style_header(ws, row, col, value):
    """Celda con estilo de encabezado."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    return cell


def write_result_sheet(ws, Z, t_ms, start_times, lb, instance_name, method_name, n):
    """
    Escribe resultados en una hoja de Excel.

    Formato Anexo 3:
        Fila 1:  Z   t_ms
        Fila 2:  s_0  s_1  ...  s_{n-1}
    (con filas adicionales de informacion y cota inferior)
    """
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    # Informacion general
    style_header(ws, 1, 1, "Instancia");    ws.cell(row=1, column=2, value=instance_name)
    style_header(ws, 2, 1, "Metodo");       ws.cell(row=2, column=2, value=method_name)
    style_header(ws, 3, 1, "Parametros")
    ws.cell(row=3, column=2, value=f"nsol={nsol}, alpha={alpha}, K={K}, r={r}")

    # Resultado principal (formato Anexo 3)
    style_header(ws, 5, 1, "Z (Flow Time)")
    style_header(ws, 5, 2, "Tiempo cómputo (ms)")
    ws.cell(row=6, column=1, value=Z)
    ws.cell(row=6, column=2, value=round(t_ms))

    # Cota inferior y gap
    style_header(ws, 8, 1, "Cota Inferior (LB)")
    style_header(ws, 8, 2, "Gap (%)")
    ws.cell(row=9, column=1, value=lb)
    gap = round((Z - lb) / lb * 100, 2) if lb > 0 else 0
    ws.cell(row=9, column=2, value=gap)

    # Tiempos de inicio (s_0, s_1, ..., s_{n-1})
    style_header(ws, 11, 1, f"Tiempos de inicio — {n} trabajos")

    COLS = 20   # trabajos por fila
    for i, s in enumerate(start_times):
        col = (i % COLS) + 1
        row = 12 + (i // COLS)
        ws.cell(row=row, column=col, value=s)

    for ci in range(1, COLS + 1):
        col_ltr = get_column_letter(ci)
        if ws.column_dimensions[col_ltr].width < 9:
            ws.column_dimensions[col_ltr].width = 9


def run_on_instance(filepath, instance_name):
    """Ejecuta los tres algoritmos sobre una instancia."""
    print(f"\n{'='*60}")
    print(f"  Instancia: {instance_name}")
    print(f"{'='*60}")

    n, m, ops, release_dates = read_nwjssp(filepath)
    print(f"  n={n} trabajos, m={m} maquinas")

    _, totals, _ = precompute(n, ops)
    lb = lower_bound(n, release_dates, totals)
    print(f"  Cota inferior (LB): {lb:,}")

    # 1. Constructivo Greedy
    print("  [1/3] Constructivo Greedy...", end=" ", flush=True)
    Z_c, S_c, st_c, ft_c, t_c = Constructive(n, m, ops, release_dates)
    gap_c = round((Z_c - lb) / lb * 100, 2) if lb > 0 else 0
    print(f"Z={Z_c:,}  gap={gap_c}%  t={round(t_c)}ms")

    # 2. GRASP
    print(f"  [2/3] GRASP (alpha={alpha}, nsol={nsol})...", end=" ", flush=True)
    Z_g, S_g, st_g, ft_g, t_g = GRASP1(n, m, ops, release_dates, alpha, nsol)
    gap_g = round((Z_g - lb) / lb * 100, 2) if lb > 0 else 0
    print(f"Z={Z_g:,}  gap={gap_g}%  t={round(t_g)}ms")

    # 3. Noise
    print(f"  [3/3] Noise (r={r}, nsol={nsol})...", end=" ", flush=True)
    Z_n, S_n, st_n, ft_n, t_n = Noise(n, m, ops, release_dates, r, nsol)
    gap_n = round((Z_n - lb) / lb * 100, 2) if lb > 0 else 0
    print(f"Z={Z_n:,}  gap={gap_n}%  t={round(t_n)}ms")

    return n, lb, (Z_c, st_c, t_c), (Z_g, st_g, t_g), (Z_n, st_n, t_n)


def build_excel(method_name, results, output_path):
    """Genera un archivo Excel con una hoja por instancia."""
    wb = Workbook()
    wb.remove(wb.active)

    for instance_name, (n, lb, Z, start_times, t_ms) in results.items():
        ws = wb.create_sheet(title=instance_name[:31])
        write_result_sheet(ws, Z, t_ms, start_times, lb, instance_name, method_name, n)

    wb.save(output_path)
    print(f"  -> Guardado: {output_path}")


def main():
    print("=" * 60)
    print("  NWJSSP — Metodos Constructivos y Aleatorizados")
    print("  Nicolas Pena — Universidad EAFIT")
    print("=" * 60)
    print(f"  Parametros: nsol={nsol}, alpha={alpha}, K={K}, r={r}")

    if not os.path.isdir(INSTANCES_FOLDER):
        print(f"\nERROR: No se encontro la carpeta '{INSTANCES_FOLDER}/'.")
        print("Crea la carpeta y coloca ahi los archivos .txt de las instancias.")
        return

    instance_files = sorted(
        f for f in os.listdir(INSTANCES_FOLDER) if f.endswith(".txt")
    )

    if not instance_files:
        print(f"\nERROR: No se encontraron archivos .txt en '{INSTANCES_FOLDER}/'.")
        return

    print(f"\n  Instancias encontradas: {len(instance_files)}")
    for f in instance_files:
        print(f"    - {f}")

    results_c = {}
    results_g = {}
    results_n = {}

    t_total_start = time.time()

    for filename in instance_files:
        instance_name = filename.replace(".txt", "")
        filepath = os.path.join(INSTANCES_FOLDER, filename)

        n, lb, res_c, res_g, res_n = run_on_instance(filepath, instance_name)

        Z_c, st_c, t_c = res_c
        Z_g, st_g, t_g = res_g
        Z_n, st_n, t_n = res_n

        results_c[instance_name] = (n, lb, Z_c, st_c, t_c)
        results_g[instance_name] = (n, lb, Z_g, st_g, t_g)
        results_n[instance_name] = (n, lb, Z_n, st_n, t_n)

    t_total = (time.time() - t_total_start) / 60

    print(f"\n{'='*60}")
    print("  Generando archivos Excel...")
    print(f"{'='*60}")

    build_excel("Constructivo Greedy",
                results_c,
                "NWJSSP_NicolasPena_Constructivo.xlsx")

    build_excel("GRASP",
                results_g,
                "NWJSSP_NicolasPena_GRASP.xlsx")

    build_excel("Noise",
                results_n,
                "NWJSSP_NicolasPena_Noise.xlsx")

    print(f"\n  Tiempo total: {t_total:.2f} minutos")
    print("  Listo.")


if __name__ == "__main__":
    main()
